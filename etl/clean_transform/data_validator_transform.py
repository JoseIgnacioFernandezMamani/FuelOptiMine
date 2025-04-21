import os
import shutil
from openpyxl import load_workbook, Workbook

def eliminar_columnas_y_formato():
    origen = os.path.join("..", "data-set", "val_data_sensor", "raw_data")
    backup = os.path.join("..", "data-set", "val_data_sensor", "backup_data")
    inicio = 2
    paso = 3

    if not os.path.exists(backup):
        os.makedirs(backup)

    for archivo in os.listdir(origen):
        if archivo.startswith('EQUIPO') and archivo.endswith('.xlsx'):
            ruta_archivo = os.path.join(origen, archivo)

            try:
                if os.path.getsize(ruta_archivo) == 0:
                    print(f"⚠️ {archivo} está vacío, se omite.")
                    continue

                # Backup
                shutil.copy(ruta_archivo, os.path.join(backup, archivo))

                # Cargar archivo original
                wb_original = load_workbook(ruta_archivo)
                wb_limpio = Workbook()
                wb_limpio.remove(wb_limpio.active)  # Quitar hoja por defecto

                for hoja_nombre in wb_original.sheetnames:
                    hoja_origen = wb_original[hoja_nombre]
                    hoja_nueva = wb_limpio.create_sheet(title=hoja_nombre)

                    for fila in hoja_origen.iter_rows(values_only=True):
                        hoja_nueva.append(list(fila))  # Solo valores

                    # Eliminar columnas en la hoja nueva
                    total_columnas = hoja_nueva.max_column
                    columnas_a_eliminar = list(range(2, total_columnas + 1, 3))
                    
                    for col in sorted(columnas_a_eliminar, reverse=True):
                        if col <= hoja_nueva.max_column:
                            hoja_nueva.delete_cols(col)

                    print(f"   🧹 Hoja '{hoja_nombre}': columnas eliminadas -> {columnas_a_eliminar}")

                # Guardar archivo limpio y reemplazar
                wb_limpio.save(ruta_archivo)
                print(f"✅ {archivo} formateado y procesado correctamente")

            except Exception as e:
                print(f"❌ Error en {archivo}: {str(e)}")

if __name__ == "__main__":
    print("=== INICIO DEL LIMPIADOR TOTAL ===")
    eliminar_columnas_y_formato()
    print("=== FINALIZADO ===")








import os
import shutil
from openpyxl import load_workbook, Workbook

def reestructurar_columnas():
    origen = os.path.join("..", "data-set", "val_data_sensor", "raw_data")
    backup = os.path.join("..", "data-set", "val_data_sensor", "backup_data")

    if not os.path.exists(backup):
        os.makedirs(backup)

    for archivo in os.listdir(origen):
        if archivo.startswith("EQUIPO") and archivo.endswith(".xlsx"):
            ruta_archivo = os.path.join(origen, archivo)

            try:
                if os.path.getsize(ruta_archivo) == 0:
                    print(f"⚠️ {archivo} está vacío. Se omite.")
                    continue

                # Backup
                shutil.copy(ruta_archivo, os.path.join(backup, archivo))

                # Cargar original
                wb_origen = load_workbook(ruta_archivo)
                wb_nuevo = Workbook()
                hoja_destino = wb_nuevo.active
                hoja_destino.title = "Unificado"

                fila_actual = 1  # Control de filas en hoja destino

                for nombre_hoja in wb_origen.sheetnames:
                    ws = wb_origen[nombre_hoja]
                    max_col = ws.max_column
                    max_row = ws.max_row

                    # Recorrer columnas de dos en dos (C-D, E-F, etc.)
                    for col in range(1, max_col + 1, 2):  # 1-based indexing
                        if col + 1 > max_col:
                            break  # si no hay par completo, saltar

                        for fila in range(1, max_row + 1):
                            val1 = ws.cell(row=fila, column=col).value
                            val2 = ws.cell(row=fila, column=col + 1).value

                            # Solo agregamos si hay al menos un valor
                            if val1 is not None or val2 is not None:
                                hoja_destino.cell(row=fila_actual, column=1, value=val1)
                                hoja_destino.cell(row=fila_actual, column=2, value=val2)
                                fila_actual += 1

                # Guardar sobre el archivo original
                wb_nuevo.save(ruta_archivo)
                print(f"✅ {archivo} reestructurado exitosamente")

            except Exception as e:
                print(f"❌ Error en {archivo}: {str(e)}")

if __name__ == "__main__":
    print("=== INICIO REESTRUCTURACIÓN DE COLUMNAS ===")
    reestructurar_columnas()
    print("=== FINALIZADO ===")
